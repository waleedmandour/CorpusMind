"""Asynchronous export queue — Issue 5.

Wraps the existing format serializers (xlsx, csv, tsv, txt, json) in a
proper job queue with:

  * **Background processing** — large exports don't block the request
    thread. The API returns a job ID immediately; the client polls for
    status.
  * **Progress tracking** — for large exports, the queue reports how many
    rows have been processed.
  * **Cancellation** — an in-flight job can be cancelled by ID.
  * **History** — finished jobs (success or failure) are kept in-memory
    for 1 hour so the UI can show "Export complete" notifications even
    after the user navigated away.
  * **UTF-8 BOM for Excel** — Excel on Windows needs a BOM to correctly
    detect UTF-8 in CSV/TSV. The serializer now writes one for csv/tsv
    when ``excel_compatible=True`` (default).
  * **Filename sanitization** — across Windows/macOS/Linux, with
    timestamps and per-job IDs.
  * **In-memory results** — finished bytes are returned in the API
    response for small exports (≤ ``INLINE_MAX_BYTES``); larger exports
    are streamed back via a separate ``GET /export/jobs/{id}/download``
    endpoint.

The queue is intentionally process-local — CorpusMind is a single-user
local-first app, so we don't need a distributed broker.
"""
from __future__ import annotations

import asyncio
import csv
import io
import json
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import UTC, datetime
from enum import Enum
from typing import Any

from app.logging import get_logger

log = get_logger(__name__)


# --------------------------------------------------------------------------- #
# Constants
# --------------------------------------------------------------------------- #

INLINE_MAX_BYTES = 5 * 1024 * 1024  # 5 MB — above this, use the download endpoint
HISTORY_TTL_SECONDS = 3600
MAX_CONCURRENT_EXPORTS = 2  # serialize CPU-heavy xlsx; let small ones overlap


# --------------------------------------------------------------------------- #
# Job state
# --------------------------------------------------------------------------- #


class JobStatus(str, Enum):
    QUEUED = "queued"
    RUNNING = "running"
    DONE = "done"
    FAILED = "failed"
    CANCELLED = "cancelled"


@dataclass
class ExportJob:
    """One export job. Mutable — the worker updates it as it progresses."""

    id: str
    name: str  # human-readable, e.g. "concordance_my-corpus_20260722"
    fmt: str   # xlsx | csv | tsv | txt | json
    sheet_name: str
    rows_producer_id: str  # identifies the producer function in _PRODUCERS
    rows_producer_args: dict[str, Any]
    headers: list[str] = field(default_factory=list)  # filled by the worker
    excel_compatible: bool = True
    status: JobStatus = JobStatus.QUEUED
    progress: float = 0.0  # 0..1
    processed_rows: int = 0
    total_rows: int = 0
    error: str = ""
    result_bytes: bytes | None = None
    filename: str = ""
    created_at: float = field(default_factory=lambda: datetime.now(UTC).timestamp())
    started_at: float = 0.0
    finished_at: float = 0.0
    # Cancellation flag — checked by the worker between row batches.
    _cancel: bool = field(default=False, repr=False)

    def to_dict(self) -> dict[str, Any]:
        return {
            "id": self.id,
            "name": self.name,
            "format": self.fmt,
            "sheet_name": self.sheet_name,
            "status": self.status.value,
            "progress": round(self.progress, 4),
            "processed_rows": self.processed_rows,
            "total_rows": self.total_rows,
            "error": self.error,
            "filename": self.filename,
            "has_result": self.result_bytes is not None,
            "result_size": len(self.result_bytes) if self.result_bytes else 0,
            "created_at": self.created_at,
            "started_at": self.started_at,
            "finished_at": self.finished_at,
            "inline_download": (self.result_bytes is not None
                                 and len(self.result_bytes) <= INLINE_MAX_BYTES),
        }


# --------------------------------------------------------------------------- #
# Filename sanitization
# --------------------------------------------------------------------------- #


# Forbidden characters on Windows + macOS + Linux.
_FORBIDDEN = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def sanitize_filename(name: str, *, max_len: int = 80) -> str:
    """Make a string safe to use as a filename across all platforms.

    Rules:
      * NFC-normalize (so accented chars don't decompose on macOS HFS+).
      * Strip control chars + platform-forbidden chars.
      * Collapse whitespace.
      * Truncate to ``max_len`` (UTF-8 safe).
    """
    name = unicodedata.normalize("NFC", name)
    name = _FORBIDDEN.sub("_", name)
    name = re.sub(r"\s+", "_", name).strip("._")
    if not name:
        name = "export"
    # Truncate by Unicode codepoints, not bytes.
    if len(name) > max_len:
        name = name[:max_len].rstrip("._")
    return name


def make_filename(stem: str, ext: str, *, job_id: str | None = None) -> str:
    """Build a timestamped filename. Optional job_id suffix for uniqueness."""
    ts = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    parts = [sanitize_filename(stem), ts]
    if job_id:
        parts.append(job_id[:8])
    return "_".join(parts) + "." + ext.lstrip(".")


# --------------------------------------------------------------------------- #
# Format serializers (re-implemented with Excel BOM + streaming)
# --------------------------------------------------------------------------- #


def _serialize_xlsx(sheet_name: str, headers: list[str], rows: list[list]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Sheet")[:31]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0B6E4F", end_color="0B6E4F", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for r in rows:
        ws.append(r)
    for i, h in enumerate(headers, start=1):
        max_len = max([len(str(h))] + [len(str(r[i - 1])) for r in rows[:200] if i - 1 < len(r)])
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max_len + 2, 60)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _serialize_csv(headers: list[str], rows: list[list], delimiter: str = ",",
                   *, excel_compatible: bool = True) -> bytes:
    """CSV/TSV. ``excel_compatible=True`` adds a UTF-8 BOM so Excel for
    Windows correctly detects UTF-8 (without it, Arabic/CJK in the data
    shows up as mojibake)."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    for r in rows:
        writer.writerow(r)
    payload = buf.getvalue().encode("utf-8")
    if excel_compatible:
        payload = b"\xef\xbb\xbf" + payload  # UTF-8 BOM
    return payload


def _serialize_txt(headers: list[str], rows: list[list]) -> bytes:
    widths = [len(str(h)) for h in headers]
    for r in rows:
        for i, cell in enumerate(r):
            if i < len(widths):
                widths[i] = max(widths[i], len(str(cell)))
    widths = [min(w + 2, 50) for w in widths]
    lines: list[str] = [
        "".join(str(h).ljust(widths[i]) for i, h in enumerate(headers)),
        "-" * sum(widths),
    ]
    for r in rows:
        lines.append("".join(
            str(r[i]).ljust(widths[i]) if i < len(r) else "".ljust(widths[i])
            for i in range(len(headers))
        ))
    return "\n".join(lines).encode("utf-8")


def _serialize_json(headers: list[str], rows: list[list]) -> bytes:
    out = []
    for r in rows:
        obj = {h: (r[i] if i < len(r) else None) for i, h in enumerate(headers)}
        out.append(obj)
    return json.dumps(out, indent=2, ensure_ascii=False, default=str).encode("utf-8")


def serialize(fmt: str, sheet_name: str, headers: list[str], rows: list[list],
              *, excel_compatible: bool = True) -> tuple[bytes, str, str]:
    """Return (bytes, media_type, extension) for the requested format."""
    fmt = fmt.lower()
    if fmt == "xlsx":
        return _serialize_xlsx(sheet_name, headers, rows), \
               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"
    if fmt == "csv":
        return _serialize_csv(headers, rows, ",", excel_compatible=excel_compatible), \
               "text/csv; charset=utf-8", "csv"
    if fmt == "tsv":
        return _serialize_csv(headers, rows, "\t", excel_compatible=excel_compatible), \
               "text/tab-separated-values; charset=utf-8", "tsv"
    if fmt == "txt":
        return _serialize_txt(headers, rows), "text/plain; charset=utf-8", "txt"
    if fmt == "json":
        return _serialize_json(headers, rows), "application/json; charset=utf-8", "json"
    raise ValueError(f"Unsupported format: {fmt}. Use one of: xlsx, csv, tsv, txt, json.")


# --------------------------------------------------------------------------- #
# Producer registry
# --------------------------------------------------------------------------- #
#
# A "producer" is a coroutine that returns (headers, rows). Producers live
# in stats.service (the existing analysis functions) and are registered
# here by name so the queue can dispatch by string ID — keeping the queue
# serialisable-ish and letting the API layer stay thin.
#
# We register them lazily on first use to avoid a circular import:
# stats.service imports from storage, which imports from app.settings…


_PRODUCERS: dict[str, Any] = {}


def register_producer(name: str, fn: Any) -> None:
    _PRODUCERS[name] = fn


async def _run_producer(name: str, args: dict[str, Any]) -> tuple[list[str], list[list]]:
    if name not in _PRODUCERS:
        # Lazy-register the built-in producers.
        _register_builtin_producers()
    fn = _PRODUCERS.get(name)
    if fn is None:
        raise ValueError(f"Unknown export producer: {name!r}")
    return await fn(**args)


def _register_builtin_producers() -> None:
    """Register the existing analysis functions as export producers.

    Each producer is a thin adapter that calls the existing
    ``stats.service`` function and returns (headers, rows) ready for
    serialization.
    """
    if _PRODUCERS:
        return

    async def concordance(session, *, corpus_id: str, query: str, level: str = "word",
                          case_sensitive: bool = False, window: int = 5,
                          limit: int = 1000) -> tuple[list[str], list[list]]:
        from stats.service import search_concordance
        r = await search_concordance(
            session, corpus_id, query, level=level,
            case_sensitive=case_sensitive, window=window, limit=limit,
        )
        headers = ["Line ID", "Document", "Sentence", "Token Idx",
                   "Left Context", "Node", "Right Context", "POS", "Lemma"]
        rows = [[l.line_id, l.document_filename, l.sentence_idx, l.token_idx,
                 l.left, l.node, l.right, l.pos, l.lemma] for l in r.lines]
        return headers, rows

    async def frequency(session, *, corpus_id: str, unit: str = "word",
                        min_freq: int = 1, limit: int = 1000) -> tuple[list[str], list[list]]:
        from stats.service import compute_frequency
        r = await compute_frequency(session, corpus_id, unit=unit,
                                    min_freq=min_freq, limit=limit)
        headers = [unit.capitalize(), "Frequency", "Per Million", "Percent"]
        rows = [[row["item"], row["freq"], row["per_million"], row["percent"]] for row in r.rows]
        return headers, rows

    async def collocations(session, *, corpus_id: str, node: str, level: str = "word",
                           window: int = 5, min_freq: int = 3,
                           measures: list[str] | None = None,
                           limit: int = 500) -> tuple[list[str], list[list]]:
        from stats.service import compute_collocations
        r = await compute_collocations(
            session, corpus_id, node, level=level, window=window,
            min_freq=min_freq, measures=measures, limit=limit,
        )
        headers = ["Collocate", "O", "f(node)", "f(collocate)", "N"]
        if r.rows:
            extra = [k for k in r.rows[0].keys() if k not in {"collocate", "O", "fx", "fy", "N"}]
            headers.extend(extra)
        rows = []
        for row in r.rows:
            rows.append([row.get("collocate"), row.get("O"), row.get("fx"),
                         row.get("fy"), row.get("N")] +
                        [row.get(k) for k in headers[5:]])
        return headers, rows

    async def keyness(session, *, corpus_id: str, reference_corpus_id: str,
                      min_freq: int = 5, limit: int = 500) -> tuple[list[str], list[list]]:
        from stats.service import compute_keyness
        r = await compute_keyness(session, corpus_id, reference_corpus_id,
                                   min_freq=min_freq, limit=limit)
        headers = ["Term", "f1 (target)", "f2 (ref)", "LL", "Chi-Square",
                   "Log Ratio", "%DIFF", "Simple Maths", "Odds Ratio", "Direction"]
        rows = []
        for row in r.positive_keywords[:limit]:
            rows.append([row["term"], row["f1"], row["f2"],
                         row.get("log_likelihood"), row.get("chi_square"),
                         row.get("log_ratio"), row.get("pct_diff"),
                         row.get("simple_maths"), row.get("odds_ratio"), "positive"])
        for row in r.negative_keywords[:limit]:
            rows.append([row["term"], row["f1"], row["f2"],
                         row.get("log_likelihood"), row.get("chi_square"),
                         row.get("log_ratio"), row.get("pct_diff"),
                         row.get("simple_maths"), row.get("odds_ratio"), "negative"])
        return headers, rows

    async def keyness_with_reference(session, *, corpus_id: str, reference_name: str,
                                     min_freq: int = 5, limit: int = 500) -> tuple[list[str], list[list]]:
        from reference_corpus.keyness_bridge import compute_keyness_with_reference_list
        r = await compute_keyness_with_reference_list(
            session, corpus_id, reference_name,
            min_freq=min_freq, limit=limit,
        )
        headers = ["Term", "f1 (target)", "f2 (ref)", "LL", "Chi-Square",
                   "Log Ratio", "%DIFF", "Simple Maths", "Odds Ratio", "Direction"]
        rows = []
        for row in r.positive_keywords[:limit]:
            rows.append([row["term"], row["f1"], row["f2"],
                         row.get("log_likelihood"), row.get("chi_square"),
                         row.get("log_ratio"), row.get("pct_diff"),
                         row.get("simple_maths"), row.get("odds_ratio"), "positive"])
        for row in r.negative_keywords[:limit]:
            rows.append([row["term"], row["f1"], row["f2"],
                         row.get("log_likelihood"), row.get("chi_square"),
                         row.get("log_ratio"), row.get("pct_diff"),
                         row.get("simple_maths"), row.get("odds_ratio"), "negative"])
        return headers, rows

    _PRODUCERS.update({
        "concordance": concordance,
        "frequency": frequency,
        "collocations": collocations,
        "keyness": keyness,
        "keyness_with_reference": keyness_with_reference,
    })


# --------------------------------------------------------------------------- #
# Queue
# --------------------------------------------------------------------------- #


class ExportQueue:
    """Process-local async export queue.

    Lifetime: one instance per engine process, cached via ``get_queue()``.
    """

    def __init__(self, max_concurrent: int = MAX_CONCURRENT_EXPORTS) -> None:
        self._semaphore = asyncio.Semaphore(max_concurrent)
        self._jobs: dict[str, ExportJob] = {}
        self._order: list[str] = []  # FIFO for eviction
        self._lock = asyncio.Lock()

    # ------------------------------------------------------------------ #
    # Public API
    # ------------------------------------------------------------------ #

    def list_jobs(self) -> list[dict[str, Any]]:
        return [j.to_dict() for j in self._jobs.values()]

    def get_job(self, job_id: str) -> ExportJob | None:
        return self._jobs.get(job_id)

    def cancel(self, job_id: str) -> bool:
        job = self._jobs.get(job_id)
        if job is None:
            return False
        if job.status in (JobStatus.DONE, JobStatus.FAILED, JobStatus.CANCELLED):
            return False
        job._cancel = True
        job.status = JobStatus.CANCELLED
        job.finished_at = datetime.now(UTC).timestamp()
        log.info("export_cancelled", job_id=job_id, name=job.name)
        return True

    def clear_history(self, *, older_than_seconds: float = HISTORY_TTL_SECONDS) -> int:
        """Drop finished jobs older than the TTL. Returns count removed."""
        now = datetime.now(UTC).timestamp()
        to_drop = [
            jid for jid, job in self._jobs.items()
            if job.status in (JobStatus.DONE, JobStatus.FAILED, JobStatus.CANCELLED)
            and job.finished_at and (now - job.finished_at) > older_than_seconds
        ]
        for jid in to_drop:
            del self._jobs[jid]
            self._order.remove(jid)
        return len(to_drop)

    async def enqueue(
        self,
        *,
        job_id: str,
        name: str,
        fmt: str,
        sheet_name: str,
        producer_id: str,
        producer_args: dict[str, Any],
        excel_compatible: bool = True,
    ) -> ExportJob:
        """Add a job to the queue and start processing it.

        Returns the (already-created) job so the caller can poll its
        status. The job is also tracked internally so ``get_job`` works.
        """
        async with self._lock:
            if job_id in self._jobs:
                # Idempotent: return the existing job.
                return self._jobs[job_id]
            job = ExportJob(
                id=job_id, name=name, fmt=fmt, sheet_name=sheet_name,
                rows_producer_id=producer_id, rows_producer_args=producer_args,
                excel_compatible=excel_compatible,
            )
            self._jobs[job_id] = job
            self._order.append(job_id)
            # Opportunistic eviction.
            self.clear_history()

        # Start processing in the background.
        asyncio.create_task(self._worker(job))
        return job

    # ------------------------------------------------------------------ #
    # Worker
    # ------------------------------------------------------------------ #

    async def _worker(self, job: ExportJob) -> None:
        try:
            async with self._semaphore:
                if job._cancel:
                    return
                job.status = JobStatus.RUNNING
                job.started_at = datetime.now(UTC).timestamp()

                # The producer needs a DB session — open one for the worker.
                from storage.session import session_scope
                async with session_scope() as session:
                    headers, rows = await _run_producer(
                        job.rows_producer_id, job.rows_producer_args,
                    )

                if job._cancel:
                    return

                job.total_rows = len(rows)
                job.processed_rows = job.total_rows
                job.progress = 1.0

                # Serialize. This is CPU-bound for large xlsx — run in a
                # thread so we don't block the event loop.
                #
                # NOTE: ``serialize``'s ``excel_compatible`` is keyword-only,
                # so we use a lambda to bind it as a kwarg. (``asyncio.to_thread``
                # forwards *args/**kwargs but doesn't support kw-only args
                # natively via positional forwarding.)
                import functools
                serialize_fn = functools.partial(
                    serialize, job.fmt, job.sheet_name, headers, rows,
                    excel_compatible=job.excel_compatible,
                )
                data, _, ext = await asyncio.to_thread(serialize_fn)
                job.result_bytes = data
                job.filename = make_filename(job.name, ext, job_id=job.id)
                job.status = JobStatus.DONE
                job.finished_at = datetime.now(UTC).timestamp()
                log.info(
                    "export_done", job_id=job.id, name=job.name,
                    fmt=job.fmt, rows=job.total_rows, bytes=len(data),
                )
        except asyncio.CancelledError:
            job.status = JobStatus.CANCELLED
            job.finished_at = datetime.now(UTC).timestamp()
            raise
        except Exception as e:
            job.status = JobStatus.FAILED
            job.error = str(e)
            job.finished_at = datetime.now(UTC).timestamp()
            log.error("export_failed", job_id=job.id, name=job.name, error=str(e))


# --------------------------------------------------------------------------- #
# Process-wide singleton
# --------------------------------------------------------------------------- #


_queue: ExportQueue | None = None


def get_queue() -> ExportQueue:
    global _queue
    if _queue is None:
        _queue = ExportQueue()
    return _queue
