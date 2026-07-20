"""Export endpoints: multi-format export for all analysis types.

Supported formats:
  - xlsx  (Excel, openpyxl) — styled, multi-row, good for spreadsheets
  - csv   (comma-separated) — universal, opens in any tool
  - tsv   (tab-separated) — good for copy-paste into Excel/Sheets
  - txt   (plain text, fixed-width) — for quick inspection / emails
  - json  (structured) — for programmatic use / re-import

Also provides:
  - POST /corpora/{cid}/export/collocations.network.svg — collocation
    network diagram as an SVG (vector, scales to any size)
  - POST /corpora/{cid}/export/collocations.network.png — same diagram
    as a PNG raster

The Methods PDF endpoint (GET /corpora/{cid}/methods.pdf) is unchanged.
"""
from __future__ import annotations

import csv
import io
import json
import urllib.parse
from datetime import UTC, datetime

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from app.logging import get_logger
from stats.service import (
    compute_collocations,
    compute_frequency,
    compute_keyness,
    search_concordance,
)
from storage.models import Corpus
from storage.session import get_session

log = get_logger(__name__)
router = APIRouter()


# --------------------------------------------------------------------------- #
# Request models
# --------------------------------------------------------------------------- #


class ExportConcordanceRequest(BaseModel):
    query: str
    level: str = "word"
    case_sensitive: bool = False
    window: int = 5
    limit: int = 1000


class ExportFrequencyRequest(BaseModel):
    unit: str = "word"
    min_freq: int = 1
    limit: int = 1000


class ExportCollocationRequest(BaseModel):
    node: str
    level: str = "word"
    window: int = 5
    min_freq: int = 3
    measures: list[str] | None = None
    limit: int = 500


class ExportKeynessRequest(BaseModel):
    reference_corpus_id: str
    min_freq: int = 5
    limit: int = 500


ExportFormat = str  # "xlsx" | "csv" | "tsv" | "txt" | "json"


# --------------------------------------------------------------------------- #
# Format serializers
# --------------------------------------------------------------------------- #


def _xlsx_bytes(sheet_name: str, headers: list[str], rows: list[list]) -> bytes:
    """Styled Excel workbook as bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0B6E4F", end_color="0B6E4F", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for r in rows:
        ws.append(r)
    for i, h in enumerate(headers, start=1):
        max_len = max([len(str(h))] + [len(str(r[i - 1])) for r in rows[:100] if i - 1 < len(r)])
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max_len + 2, 60)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _csv_bytes(headers: list[str], rows: list[list], delimiter: str = ",") -> bytes:
    """CSV or TSV as bytes."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    for r in rows:
        writer.writerow(r)
    return buf.getvalue().encode("utf-8")


def _txt_bytes(headers: list[str], rows: list[list]) -> bytes:
    """Fixed-width plain text table."""
    # Compute column widths
    widths = [len(str(h)) for h in headers]
    for r in rows:
        for i, cell in enumerate(r):
            if i < len(widths):
                widths[i] = max(widths[i], len(str(cell)))
    widths = [min(w + 2, 50) for w in widths]  # cap at 50

    lines: list[str] = []
    # Header
    lines.append("".join(str(h).ljust(widths[i]) for i, h in enumerate(headers)))
    # Separator
    lines.append("-" * sum(widths))
    # Rows
    for r in rows:
        lines.append("".join(str(r[i]).ljust(widths[i]) if i < len(r) else "".ljust(widths[i]) for i in range(len(headers))))
    return "\n".join(lines).encode("utf-8")


def _json_bytes(headers: list[str], rows: list[list]) -> bytes:
    """Array of objects (each row → {header: value})."""
    out = []
    for r in rows:
        obj = {}
        for i, h in enumerate(headers):
            obj[h] = r[i] if i < len(r) else None
        out.append(obj)
    return json.dumps(out, indent=2, ensure_ascii=False, default=str).encode("utf-8")


def _serialize(format: ExportFormat, sheet_name: str, headers: list[str], rows: list[list]) -> tuple[bytes, str, str]:
    """Return (data, media_type, extension) for the requested format."""
    if format == "xlsx":
        return _xlsx_bytes(sheet_name, headers, rows), \
               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"
    if format == "csv":
        return _csv_bytes(headers, rows, ","), "text/csv; charset=utf-8", "csv"
    if format == "tsv":
        return _csv_bytes(headers, rows, "\t"), "text/tab-separated-values; charset=utf-8", "tsv"
    if format == "txt":
        return _txt_bytes(headers, rows), "text/plain; charset=utf-8", "txt"
    if format == "json":
        return _json_bytes(headers, rows), "application/json; charset=utf-8", "json"
    raise HTTPException(400, f"Unsupported format: {format}. Use one of: xlsx, csv, tsv, txt, json.")


def _make_response(data: bytes, media_type: str, filename: str) -> StreamingResponse:
    """Build a StreamingResponse with a safe Content-Disposition header."""
    # RFC 5987 encoding for non-ASCII filenames
    safe_fname = filename.encode("ascii", "replace").decode("ascii")
    encoded_fname = urllib.parse.quote(filename)
    return StreamingResponse(
        io.BytesIO(data),
        media_type=media_type,
        headers={
            "Content-Disposition": f"attachment; filename=\"{safe_fname}\"; filename*=UTF-8''{encoded_fname}",
        },
    )


# --------------------------------------------------------------------------- #
# Backwards-compatible xlsx endpoints (kept so old URLs don't break)
# --------------------------------------------------------------------------- #


@router.post("/corpora/{cid}/export/concordance.xlsx")
async def export_concordance_xlsx(cid: str, body: ExportConcordanceRequest, session: AsyncSession = Depends(get_session)) -> StreamingResponse:
    return await _export_concordance(cid, body, "xlsx", session)


@router.post("/corpora/{cid}/export/frequency.xlsx")
async def export_frequency_xlsx(cid: str, body: ExportFrequencyRequest, session: AsyncSession = Depends(get_session)) -> StreamingResponse:
    return await _export_frequency(cid, body, "xlsx", session)


@router.post("/corpora/{cid}/export/collocations.xlsx")
async def export_collocations_xlsx(cid: str, body: ExportCollocationRequest, session: AsyncSession = Depends(get_session)) -> StreamingResponse:
    return await _export_collocations(cid, body, "xlsx", session)


@router.post("/corpora/{cid}/export/keyness.xlsx")
async def export_keyness_xlsx(cid: str, body: ExportKeynessRequest, session: AsyncSession = Depends(get_session)) -> StreamingResponse:
    return await _export_keyness(cid, body, "xlsx", session)


# --------------------------------------------------------------------------- #
# Multi-format endpoints (the new, clean API)
# --------------------------------------------------------------------------- #


async def _export_concordance(cid: str, body: ExportConcordanceRequest, fmt: ExportFormat, session: AsyncSession) -> StreamingResponse:
    if not await session.get(Corpus, cid):
        raise HTTPException(404, "Corpus not found")
    r = await search_concordance(
        session, cid, body.query, level=body.level,
        case_sensitive=body.case_sensitive, window=body.window, limit=body.limit,
    )
    headers = ["Line ID", "Document", "Sentence", "Token Idx", "Left Context", "Node", "Right Context", "POS", "Lemma"]
    rows = [[l.line_id, l.document_filename, l.sentence_idx, l.token_idx, l.left, l.node, l.right, l.pos, l.lemma] for l in r.lines]
    data, media, ext = _serialize(fmt, "Concordance", headers, rows)
    fname = f"concordance_{cid}_{datetime.now(UTC).strftime('%Y%m%d_%H%M%S')}.{ext}"
    return _make_response(data, media, fname)


async def _export_frequency(cid: str, body: ExportFrequencyRequest, fmt: ExportFormat, session: AsyncSession) -> StreamingResponse:
    if not await session.get(Corpus, cid):
        raise HTTPException(404, "Corpus not found")
    r = await compute_frequency(session, cid, unit=body.unit, min_freq=body.min_freq, limit=body.limit)
    headers = [body.unit.capitalize(), "Frequency", "Per Million", "Percent"]
    rows = [[row["item"], row["freq"], row["per_million"], row["percent"]] for row in r.rows]
    data, media, ext = _serialize(fmt, f"Frequency ({body.unit})", headers, rows)
    fname = f"frequency_{cid}_{body.unit}_{datetime.now(UTC).strftime('%Y%m%d_%H%M%S')}.{ext}"
    return _make_response(data, media, fname)


async def _export_collocations(cid: str, body: ExportCollocationRequest, fmt: ExportFormat, session: AsyncSession) -> StreamingResponse:
    if not await session.get(Corpus, cid):
        raise HTTPException(404, "Corpus not found")
    r = await compute_collocations(
        session, cid, body.node, level=body.level, window=body.window,
        min_freq=body.min_freq, measures=body.measures, limit=body.limit,
    )
    headers = ["Collocate", "O", "f(node)", "f(collocate)", "N"]
    if r.rows:
        extra_keys = [k for k in r.rows[0].keys() if k not in {"collocate", "O", "fx", "fy", "N"}]
        headers.extend(extra_keys)
    rows = []
    for row in r.rows:
        rows.append([row.get("collocate"), row.get("O"), row.get("fx"), row.get("fy"), row.get("N")] +
                    [row.get(k) for k in headers[5:]])
    data, media, ext = _serialize(fmt, f"Collocations ({body.node})", headers, rows)
    fname = f"collocations_{cid}_{body.node}_{datetime.now(UTC).strftime('%Y%m%d_%H%M%S')}.{ext}"
    return _make_response(data, media, fname)


async def _export_keyness(cid: str, body: ExportKeynessRequest, fmt: ExportFormat, session: AsyncSession) -> StreamingResponse:
    if not await session.get(Corpus, cid):
        raise HTTPException(404, "Target corpus not found")
    r = await compute_keyness(session, cid, body.reference_corpus_id, min_freq=body.min_freq, limit=body.limit)
    headers = ["Term", "f1 (target)", "f2 (ref)", "LL", "Chi-Square", "Log Ratio", "%DIFF", "Simple Maths", "Odds Ratio", "Direction"]
    rows = []
    for row in r.positive_keywords[:body.limit]:
        rows.append([row["term"], row["f1"], row["f2"],
                     row.get("log_likelihood"), row.get("chi_square"),
                     row.get("log_ratio"), row.get("pct_diff"),
                     row.get("simple_maths"), row.get("odds_ratio"), "positive"])
    for row in r.negative_keywords[:body.limit]:
        rows.append([row["term"], row["f1"], row["f2"],
                     row.get("log_likelihood"), row.get("chi_square"),
                     row.get("log_ratio"), row.get("pct_diff"),
                     row.get("simple_maths"), row.get("odds_ratio"), "negative"])
    data, media, ext = _serialize(fmt, "Keyness", headers, rows)
    fname = f"keyness_{cid}_vs_{body.reference_corpus_id}_{datetime.now(UTC).strftime('%Y%m%d_%H%M%S')}.{ext}"
    return _make_response(data, media, fname)


@router.post("/corpora/{cid}/export/concordance")
async def export_concordance_multi(
    cid: str,
    body: ExportConcordanceRequest,
    fmt: ExportFormat = Query("xlsx", pattern="^(xlsx|csv|tsv|txt|json)$"),
    session: AsyncSession = Depends(get_session),
) -> StreamingResponse:
    """Export concordance results in the requested format."""
    return await _export_concordance(cid, body, fmt, session)


@router.post("/corpora/{cid}/export/frequency")
async def export_frequency_multi(
    cid: str,
    body: ExportFrequencyRequest,
    fmt: ExportFormat = Query("xlsx", pattern="^(xlsx|csv|tsv|txt|json)$"),
    session: AsyncSession = Depends(get_session),
) -> StreamingResponse:
    """Export frequency results in the requested format."""
    return await _export_frequency(cid, body, fmt, session)


@router.post("/corpora/{cid}/export/collocations")
async def export_collocations_multi(
    cid: str,
    body: ExportCollocationRequest,
    fmt: ExportFormat = Query("xlsx", pattern="^(xlsx|csv|tsv|txt|json)$"),
    session: AsyncSession = Depends(get_session),
) -> StreamingResponse:
    """Export collocation results in the requested format."""
    return await _export_collocations(cid, body, fmt, session)


@router.post("/corpora/{cid}/export/keyness")
async def export_keyness_multi(
    cid: str,
    body: ExportKeynessRequest,
    fmt: ExportFormat = Query("xlsx", pattern="^(xlsx|csv|tsv|txt|json)$"),
    session: AsyncSession = Depends(get_session),
) -> StreamingResponse:
    """Export keyness results in the requested format."""
    return await _export_keyness(cid, body, fmt, session)


# --------------------------------------------------------------------------- #
# Collocation network diagram (SVG + PNG)
# --------------------------------------------------------------------------- #


def _build_collocation_network_svg(node: str, rows: list[dict], max_nodes: int = 20) -> str:
    """Build an SVG collocation network diagram.

    The node word sits in the center; each collocate is placed in a circle
    around it. Edge thickness is proportional to the association strength
    (using the first available measure — usually MI or log-likelihood).
    Node radius is proportional to the collocate's raw frequency (O).
    """
    if not rows:
        return '<svg xmlns="http://www.w3.org/2000/svg" width="600" height="200"><text x="300" y="100" text-anchor="middle" fill="#888">No collocates to display.</text></svg>'

    # Take top N collocates by the first measure
    top = rows[:max_nodes]
    # Find the measure column (first non-structural key)
    measure_keys = [k for k in top[0].keys() if k not in {"collocate", "O", "fx", "fy", "N"}]
    measure_key = measure_keys[0] if measure_keys else "O"

    # Scale values
    max_measure = max(abs(float(r.get(measure_key, 0) or 0)) for r in top) or 1
    max_freq = max(float(r.get("O", 0) or 0) for r in top) or 1

    # Layout: center node + circular placement
    width, height = 800, 600
    cx, cy = width // 2, height // 2
    center_radius = 40
    radius = min(width, height) // 2 - 80

    import math
    parts: list[str] = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" '
        f'viewBox="0 0 {width} {height}" font-family="Inter, Segoe UI, sans-serif">',
        '<rect width="100%" height="100%" fill="white"/>',
        # Title
        f'<text x="{width//2}" y="30" text-anchor="middle" font-size="16" font-weight="bold" fill="#1c1f1d">'
        f'Collocation network: "{node}" (top {len(top)} by {measure_key})'
        f'</text>',
    ]

    # Edges (drawn first so nodes sit on top)
    for i, r in enumerate(top):
        angle = 2 * math.pi * i / len(top) - math.pi / 2
        x = cx + radius * math.cos(angle)
        y = cy + radius * math.sin(angle)
        measure_val = float(r.get(measure_key, 0) or 0)
        edge_width = max(0.5, 4 * abs(measure_val) / abs(max_measure))
        parts.append(
            f'<line x1="{cx}" y1="{cy}" x2="{x:.1f}" y2="{y:.1f}" '
            f'stroke="#0b6e4f" stroke-width="{edge_width:.2f}" stroke-opacity="0.5"/>'
        )

    # Collocate nodes
    for i, r in enumerate(top):
        angle = 2 * math.pi * i / len(top) - math.pi / 2
        x = cx + radius * math.cos(angle)
        y = cy + radius * math.sin(angle)
        freq = float(r.get("O", 0) or 0)
        node_r = 12 + 18 * (freq / max_freq)
        collocate = str(r.get("collocate", "?"))
        # Escape XML
        collocate_esc = collocate.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")
        parts.append(
            f'<circle cx="{x:.1f}" cy="{y:.1f}" r="{node_r:.1f}" fill="#2a9070" stroke="#0b6e4f" stroke-width="1.5"/>'
        )
        parts.append(
            f'<text x="{x:.1f}" y="{y + node_r + 14:.1f}" text-anchor="middle" font-size="11" fill="#1c1f1d">{collocate_esc}</text>'
        )

    # Center node (the node word)
    node_esc = node.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")
    parts.append(f'<circle cx="{cx}" cy="{cy}" r="{center_radius}" fill="#0b6e4f" stroke="#095c41" stroke-width="2"/>')
    parts.append(
        f'<text x="{cx}" y="{cy + 5}" text-anchor="middle" font-size="14" font-weight="bold" fill="white">{node_esc}</text>'
    )

    # Footer with parameters
    parts.append(
        f'<text x="20" y="{height - 20}" font-size="10" fill="#888">'
        f'Edge thickness = |{measure_key}|; node size = raw frequency (O). '
        f'Generated by CorpusMind.'
        f'</text>'
    )
    parts.append("</svg>")
    return "\n".join(parts)


def _svg_to_png(svg: str, width: int = 1600, height: int = 1200) -> bytes:
    """Rasterize an SVG to PNG using cairosvg (pure-Python, no system deps).

    Falls back to a clear error message if cairosvg isn't installed.
    """
    try:
        import cairosvg  # type: ignore[import-untyped]
    except ImportError as e:
        raise HTTPException(
            500,
            "PNG export requires the 'cairosvg' package. Install it with: pip install cairosvg"
        ) from e
    return cairosvg.svg2png(bytestring=svg.encode("utf-8"), output_width=width, output_height=height)


@router.post("/corpora/{cid}/export/collocations.network.svg")
async def export_collocation_network_svg(
    cid: str,
    body: ExportCollocationRequest,
    session: AsyncSession = Depends(get_session),
) -> StreamingResponse:
    """Export the collocation network as an SVG vector diagram.

    The SVG scales to any size without quality loss — good for papers,
    posters, and slides. Open it in a browser, Inkscape, or Adobe Illustrator.
    """
    if not await session.get(Corpus, cid):
        raise HTTPException(404, "Corpus not found")
    r = await compute_collocations(
        session, cid, body.node, level=body.level, window=body.window,
        min_freq=body.min_freq, measures=body.measures, limit=body.limit,
    )
    svg = _build_collocation_network_svg(body.node, r.rows)
    fname = f"collocation_network_{cid}_{body.node}_{datetime.now(UTC).strftime('%Y%m%d_%H%M%S')}.svg"
    return _make_response(svg.encode("utf-8"), "image/svg+xml; charset=utf-8", fname)


@router.post("/corpora/{cid}/export/collocations.network.png")
async def export_collocation_network_png(
    cid: str,
    body: ExportCollocationRequest,
    session: AsyncSession = Depends(get_session),
) -> StreamingResponse:
    """Export the collocation network as a PNG raster (1600x1200).

    PNG is good for embedding in Word documents, slides, or social media
    where SVG isn't supported. Requires cairosvg (install: pip install cairosvg).
    """
    if not await session.get(Corpus, cid):
        raise HTTPException(404, "Corpus not found")
    r = await compute_collocations(
        session, cid, body.node, level=body.level, window=body.window,
        min_freq=body.min_freq, measures=body.measures, limit=body.limit,
    )
    svg = _build_collocation_network_svg(body.node, r.rows)
    png = _svg_to_png(svg)
    fname = f"collocation_network_{cid}_{body.node}_{datetime.now(UTC).strftime('%Y%m%d_%H%M%S')}.png"
    return _make_response(png, "image/png", fname)


# --------------------------------------------------------------------------- #
# PDF (reportlab) — methods section auto-draft (§8.23)
# --------------------------------------------------------------------------- #


@router.get("/corpora/{cid}/methods.pdf")
async def export_methods_pdf(cid: str, session: AsyncSession = Depends(get_session)) -> StreamingResponse:
    """Auto-draft a methodology paragraph (§8.23) for the corpus's pipeline recipe."""
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    c = await session.get(Corpus, cid)
    if not c:
        raise HTTPException(404, "Corpus not found")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], textColor=colors.HexColor("#0B6E4F"))
    body = ParagraphStyle("Body", parent=styles["BodyText"], alignment=TA_LEFT, leading=15)
    recipe = c.pipeline_recipe or {}
    stats = c.stats or {}

    flow = [
        Paragraph("Methods Section (auto-drafted)", h1),
        Spacer(1, 12),
        Paragraph(
            f"The analysis was conducted using <b>CorpusMind</b> v0.1.14. "
            f"The corpus &ldquo;{c.name}&rdquo; ({c.language}) contained "
            f"{stats.get('document_count', 0)} documents, "
            f"{stats.get('token_count', 0)} tokens, and "
            f"{stats.get('type_count', 0)} types. "
            f"Texts were tokenized and annotated using "
            f"<b>{recipe.get('backend', 'spacy')}</b> with the "
            f"<b>{recipe.get('model_name', 'en_core_web_sm')}</b> model "
            f"(version {recipe.get('model_version', 'unknown')}; "
            f"spaCy {recipe.get('spacy_version', 'unknown')}). "
            f"Annotation versions are pinned per corpus for reproducibility.",
            body,
        ),
        Spacer(1, 18),
        Paragraph("Statistical measures", h1),
        Spacer(1, 6),
        Paragraph(
            "Collocation strength was computed using Mutual Information (Church &amp; Hanks, 1990), "
            "T-score, log-likelihood (Dunning, 1993), Dice, LogDice (Rychlý, 2008), chi-square, "
            "and Delta P (Gries, 2013; Ellis, 2007). Keyness was computed using log-likelihood "
            "and chi-square as significance tests, with Log Ratio (Hardie, 2014), %DIFF "
            "(Gabrielatos &amp; Marchi, 2012), Simple Maths (Kilgarriff, 2009), and Odds Ratio "
            "as effect-size measures. Dispersion was computed using Juilland's D and Gries' DP "
            "(Gries, 2008). Standardized type-token ratio (STTR) was computed over 1000-token "
            "chunks. All formulas are documented in <i>docs/METHODOLOGY.md</i>.",
            body,
        ),
        Spacer(1, 18),
        Paragraph("Reproducibility", h1),
        Spacer(1, 6),
        Paragraph(
            "Every analysis result is traceable to the exact annotation version, model, and "
            "formula version that produced it. The pipeline recipe recorded for this corpus is: "
            f"{recipe}.",
            body,
        ),
        Spacer(1, 18),
        Paragraph("AI Usage Disclosure", h1),
        Spacer(1, 6),
        Paragraph(
            "AI assistance was available via the CorpusMind AI Assistant during this analysis. "
            "All statistical results (frequency, collocation, keyness, dispersion, etc.) were "
            "computed by deterministic, pure mathematical functions &mdash; the AI model never "
            "computed a statistic. The AI Assistant's role was limited to interpreting "
            "pre-computed results and answering questions about the corpus using a grounded, "
            "citation-enforced protocol. Every AI claim is either backed by a cited tool call "
            "with a stable evidence ID (grounded) or visibly flagged as ungrounded. "
            "AI interpretations are stochastic; re-running the same query may produce different "
            "text. The specific provider, model, and number of AI turns should be reported "
            "in your manuscript's Methods section. The full AI usage audit trail is available "
            "in the engine's conversation persistence layer and can be exported via the "
            "AI Usage Disclosure endpoint.",
            body,
        ),
        Spacer(1, 18),
        Paragraph("Human Verification", h1),
        Spacer(1, 6),
        Paragraph(
            "AI-generated interpretations can be verified, rejected, or edited by the human "
            "researcher. The verification state is recorded in the audit trail alongside each "
            "AI turn. Only interpretations marked as &ldquo;accepted&rdquo; should be cited in "
            "published work.",
            body,
        ),
    ]
    doc.build(flow)
    data = buf.getvalue()
    fname = f"methods_{cid}_{datetime.now(UTC).strftime('%Y%m%d_%H%M%S')}.pdf"
    return _make_response(data, "application/pdf", fname)
